import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Consolidador EERR SAP", layout="wide")
st.title("📊 Consolidador de Reportes SAP (RFBILA00)")
st.write("Carga los reportes en los meses correspondientes. El sistema cruzará la información para generar el Estado de Resultados mensual (aislado) y acumulado (YTD).")

# --- 1. CONFIGURACIÓN DE ESTILOS EXCEL ---
def fill(hex_c): return PatternFill("solid", fgColor=hex_c)
thin = Side(style='thin', color='AAAAAA')
def bdr(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def hdr_font(sz=10, bold=True, color="FFFFFF"): return Font(bold=bold, size=sz, color=color)

C_TITLE = "1F3864"; C_HDR_BLU = "2E75B6"; C_PALE = "DEEAF1"
C_GREEN = "375623"; C_RED_HDR = "833C00"; C_GASTO = "1F3864"; C_VIOLET = "5E3292"
C_GREY1 = "F2F2F2"; C_WHITE = "FFFFFF"; C_SUB_TOT = "BDD7EE"; C_GRAND = "1F3864"

DIVS_SHOW = ['10','30','40','50','60','70','99','SIN_DIV']
DIV_LABEL = {'10':'Div.10','30':'Div.30','40':'Div.40','50':'Div.50',
             '60':'Div.60','70':'Div.70','99':'Div.99\n(Gral)','SIN_DIV':'Sin\nDiv.'}

MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

EERR_ROWS = [
    ('── INGRESOS ──', None, 1, 'section', C_GREEN),
    ('Ventas Nacionales y Export.', ['0410101'], -1, 'line', None),
    ('Servicios', ['0410105'], -1, 'line', None),
    ('TOTAL INGRESOS', '__ing__', 1, 'subtotal', C_GREEN),
    ('── COSTO DE VENTAS ──', None, 1, 'section', C_RED_HDR),
    ('Costo de Mercaderías', ['0510101'], 1, 'line', None),
    ('Estimaciones de Inventario', ['0510203'], 1, 'line', None),
    ('TOTAL COSTO DE VENTAS', '__costo__', 1, 'subtotal', C_RED_HDR),
    ('MARGEN BRUTO', '__margen__', 1, 'result', C_HDR_BLU),
    ('── GASTOS OPERACIONALES ──', None, 1, 'section', C_GASTO),
    ('Remuneraciones', ['0610101'], 1, 'line', None),
    ('Asesoría Profesional', ['0610201'], 1, 'line', None),
    ('Arriendos Bodegas', ['0610301'], 1, 'line', None),
    ('Arriendos Oficina', ['0610401'], 1, 'line', None),
    ('Fletes', ['0610501'], 1, 'line', None),
    ('Gastos de Viajes', ['0610801'], 1, 'line', None),
    ('Seg. y Gastos de Bodegas', ['0611001'], 1, 'line', None),
    ('Servicios y Gastos Varios', ['0611101'], 1, 'line', None),
    ('Suscripciones', ['0611201'], 1, 'line', None),
    ('Seguros', ['0611301'], 1, 'line', None),
    ('Materiales de Oficina', ['0611601'], 1, 'line', None),
    ('Servicios TI y Comunicaciones', ['0611701'], 1, 'line', None),
    ('Estim. Deudores Incobrables', ['0620101'], 1, 'line', None),
    ('TOTAL GASTOS OPERACIONALES', '__gto__', 1, 'subtotal', C_GASTO),
    ('RESULTADO OPERACIONAL', '__rop__', 1, 'result', C_HDR_BLU),
    ('── RESULT. NO OPERACIONAL ──', None, 1, 'section', C_VIOLET),
    ('Ingresos/Egresos No Operac.', ['071', '072'], -1, 'line', None),
    ('Diferencias de Cambio', ['081'], -1, 'line', None),
    ('TOTAL RESULT. NO OPERACIONAL', '__nop__', 1, 'subtotal', C_VIOLET),
    ('── IMPUESTOS ──', None, 1, 'section', "404040"),
    ('Impuesto a la Renta', ['091'], -1, 'line', None),
    ('RESULTADO NETO', '__neto__', 1, 'result', C_GRAND),
]

# --- 2. LÓGICA DE EXTRACCIÓN ---
def parse_num(val):
    if pd.isna(val) or str(val).strip() in ['nan', '']: return None
    s = str(val).strip().replace('.', '').replace(',', '.')
    if '*' in s or s in ['-', '']: return None
    try: return float(s)
    except: return None

def procesar_archivo_sap(file):
    df_raw = pd.read_excel(file, sheet_name='Data', header=None)
    records = []
    for i, row in df_raw.iterrows():
        soc = str(row[1]).strip() if pd.notna(row[1]) else ''
        div = str(row[2]).strip() if pd.notna(row[2]) else 'SIN_DIV'
        cuenta = str(row[3]).strip() if pd.notna(row[3]) else ''
        
        if soc == '2000' and cuenta and cuenta[0].isdigit():
            val_act = parse_num(row[9])
            val_comp = parse_num(row[11])
            if val_act is not None or val_comp is not None:
                records.append({
                    'div': div, 'cuenta': cuenta,
                    'val_act': val_act or 0.0, 'val_comp': val_comp or 0.0,
                    'val_inc': (val_act or 0.0) - (val_comp or 0.0)
                })
    return pd.DataFrame(records)

# --- 3. CONSTRUCCIÓN DEL EXCEL ---
def sum_pfx(df, div, prefixes, col):
    if prefixes is None: return 0.0
    sub = df if div=='TOTAL' else df[df['div']==div]
    mask = sub['cuenta'].apply(lambda x: any(x.startswith(p) for p in prefixes))
    return sub[mask][col].sum()

def comp_div(df, div, col):
    acc = {}
    acc['ing']   = sum_pfx(df, div, ['0410101','0410105'], col) * -1
    acc['costo'] = sum_pfx(df, div, ['0510101','0510203'], col)
    acc['gto']   = sum_pfx(df, div, ['0610101','0610201','0610301','0610401','0610501','0610801','0611001','0611101','0611201','0611301','0611601','0611701','0620101'], col)
    acc['nop']   = sum_pfx(df, div, ['071', '072', '081'], col) * -1
    acc['imp']   = sum_pfx(df, div, ['091'], col) * -1
    acc['margen']= acc['ing'] - acc['costo']
    acc['rop']   = acc['margen'] - acc['gto']
    acc['neto']  = acc['rop'] + acc['nop'] + acc['imp']
    return acc

def write_eerr_sheet(wb, df_src, col_data, sheet_name, period_label, is_first=False):
    ws = wb.active if is_first else wb.create_sheet(sheet_name)
    ws.title = sheet_name
    pre = {d: comp_div(df_src, d, col_data) for d in DIVS_SHOW + ['TOTAL']}
    ALL_COLS = DIVS_SHOW + ['TOTAL']
    N = len(ALL_COLS)
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N+1)
    c = ws.cell(1,1, "ESTADO DE RESULTADOS POR DIVISIÓN — SOCIEDAD 2000")
    c.font = hdr_font(13); c.fill = fill(C_TITLE)
    c.alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[1].height=26
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N+1)
    c = ws.cell(2,1, f"{period_label}  |  Moneda: USD")
    c.font = Font(italic=True, size=10, color=C_TITLE); c.fill = fill(C_PALE)
    c.alignment = Alignment(horizontal='center'); ws.row_dimensions[2].height=18
    
    ws.cell(3,1,"CONCEPTO").font = hdr_font(); ws.cell(3,1).fill = fill(C_HDR_BLU); ws.cell(3,1).border = bdr()
    for ci, div in enumerate(ALL_COLS, start=2):
        lbl = DIV_LABEL.get(div, div) if div != 'TOTAL' else 'TOTAL\nSOCIEDAD'
        c = ws.cell(3, ci, lbl)
        c.font = hdr_font(9); c.fill = fill(C_HDR_BLU if div!='TOTAL' else C_TITLE)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); c.border = bdr()
    ws.row_dimensions[3].height = 34
    
    cur_row = 4; alt = False
    for label, prefixes, sign, row_type, sec_color in EERR_ROWS:
        if row_type == 'section':
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=N+1)
            c = ws.cell(cur_row, 1, label)
            c.font = hdr_font(10); c.fill = fill(sec_color)
            c.alignment = Alignment(horizontal='left', indent=1); c.border = bdr()
            ws.row_dimensions[cur_row].height = 18; cur_row += 1; alt = False
            continue
        
        def get_v(div, prefixes, sign):
            p = pre[div]
            if prefixes == '__ing__':   return p['ing']
            if prefixes == '__costo__': return p['costo']
            if prefixes == '__margen__':return p['margen']
            if prefixes == '__gto__':   return p['gto']
            if prefixes == '__rop__':   return p['rop']
            if prefixes == '__nop__':   return p['nop']
            if prefixes == '__neto__':  return p['neto']
            return sum_pfx(df_src, div, prefixes, col_data) * sign
        
        row_bg = {'line': C_GREY1 if alt else C_WHITE, 'subtotal': C_SUB_TOT, 'result': sec_color or C_HDR_BLU}[row_type]
        row_fc = "FFFFFF" if row_type == 'result' else "000000"
        is_bold = row_type in ('subtotal','result')
        
        c = ws.cell(cur_row, 1, label)
        c.font = Font(bold=is_bold, size=9 if row_type=='line' else 10, color=row_fc)
        c.fill = fill(row_bg); c.alignment = Alignment(horizontal='left', indent=2 if row_type=='line' else 1); c.border = bdr()
        
        for ci, div in enumerate(ALL_COLS, start=2):
            v = get_v(div, prefixes, sign)
            bg = C_PALE if div=='TOTAL' and row_type=='line' else row_bg
            c = ws.cell(cur_row, ci, v); c.number_format = '#,##0.00'
            c.font = Font(bold=is_bold, size=9 if row_type=='line' else 10, color=row_fc)
            c.fill = fill(bg); c.alignment = Alignment(horizontal='right'); c.border = bdr()
            if row_type in ('result','subtotal') and isinstance(v, float) and v < 0:
                c.font = Font(bold=is_bold, size=9 if row_type=='line' else 10, color="FF0000" if row_type!='result' else "FFAAAA")
        
        ws.row_dimensions[cur_row].height = 16 if row_type=='line' else 18
        cur_row += 1
        if row_type == 'line': alt = not alt
    
    ws.column_dimensions['A'].width = 32
    for ci in range(2, N+2): ws.column_dimensions[get_column_letter(ci)].width = 14 if ci < N+1 else 16
    ws.freeze_panes = 'B4'

# --- 4. INTERFAZ DE STREAMLIT ---
st.subheader("📥 Carga de Archivos Mensuales")

# Crear una cuadrícula de 4 columnas para organizar los 12 meses
cols = st.columns(4)
uploads = {}

for i, mes in enumerate(MESES):
    with cols[i % 4]:
        file = st.file_uploader(f"Reporte {mes}", type=["xlsx", "xls"], key=f"file_{i}")
        if file:
            uploads[i + 1] = (mes, file)

st.divider()

if uploads:
    if st.button("⚙️ Procesar y Generar Excel Consolidado", type="primary"):
        with st.spinner("Analizando información y cruzando períodos..."):
            
            sheets_to_generate = []
            is_first_file = True
            
            # Ordenar las subidas por número de mes para procesar cronológicamente
            for m_idx in sorted(uploads.keys()):
                month_name, file = uploads[m_idx]
                df = procesar_archivo_sap(file)
                
                # Si es el primer archivo subido, y no es Enero, podemos rescatar el mes base anterior
                if is_first_file and m_idx > 1:
                    prev_month_name = MESES[m_idx - 2]
                    sheets_to_generate.append((df, 'val_comp', f"EERR {prev_month_name}", f"Base Aislado: {prev_month_name} 2026"))
                
                is_first_file = False
                
                # Agregar Pestaña Incremental (Mes Aislado)
                if m_idx == 1:
                    # Para Enero, el acumulado es igual al incremental
                    sheets_to_generate.append((df, 'val_act', f"EERR {month_name}", f"Mes Aislado: {month_name} 2026"))
                else:
                    sheets_to_generate.append((df, 'val_inc', f"EERR {month_name}", f"Mes Aislado: {month_name} 2026"))
                
                # Agregar Pestaña Acumulada YTD
                sheets_to_generate.append((df, 'val_act', f"Acumulado {month_name}", f"Acumulado YTD a {month_name} 2026"))

            # Crear el Workbook
            wb = Workbook()
            is_first_sheet = True
            
            # Generar las hojas en orden inverso (para que el Acumulado más reciente quede de primero)
            for df, col_data, sheet_name, period_label in reversed(sheets_to_generate):
                write_eerr_sheet(wb, df, col_data, sheet_name, period_label, is_first=is_first_sheet)
                is_first_sheet = False
            
            # Guardar Excel en memoria
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            st.success("✅ ¡Planilla dinámica generada con éxito!")
            st.download_button(
                label="📥 Descargar Planilla Excel",
                data=buffer,
                file_name="EERR_Consolidado_SAP.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
else:
    st.info("☝️ Sube al menos un mes para habilitar el botón de generación.")